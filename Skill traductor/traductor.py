codigo = '''
#SKILL_DATA
skpy.skill_config(name="Traductor académico",
                  description="Este es un traductor de inglés a español en el ámbito académico",
                  version="1",
                  url_image="https://tecgpt0grl0prod0stg.blob.core.windows.net/imagenes-modelos/MOD_POST_REDES_SOCIALES.svg")

#DEFINE_THE_MODEL
m_gpt_4 = model(creativity = '0.5', name = 'gpt3.5')

#DEFINE_THE_INPUTS
input = [
    {'nombre_materia': 'Escribe el nombre de la materia'},
    {'disciplina': 'Escribe la disciplina de la materia'},
    {'intencion': 'Escribe la intención de la materia'},
    {'objetivo_general': 'Escribe el objetivo general de la materia'},
    {'temas': 'Escribe los temas de la materia'},
    {'metodologia_enseñanza': 'Escribe la metodología de la enseñanza de la materia'},
    {'tiempo_estimado': 'Escribe el tiempo estimado de la materia'},
    {'politica_evaluacion': 'Escribe la política de evaluación de la materia'},
    {'perfil_profesor': 'Escribe el perfil del profesor de la materia'},
    {'requisitos': 'Escribe los requisitos de la materia'},
    {'descripcion': 'Escribe la descripción de la materia'}
]

#VARIABLES_TO_USE
personalidad_modelo = 'traductor de lenguas, especializado en el ámbito académico'
idioma_origen = 'español'
idioma_destino = 'inglés'

ordenes_nombre_materia = f'Funge el rol de un {personalidad_modelo}. Ahora, traduce de {idioma_origen} a {idioma_destino} el siguiente contenido (si no te proporciono nada después de los dos puntos entonces imprime una x): {nombre_materia}.'
ordenes_disciplina = f'Funge el rol de un {personalidad_modelo}. Ahora, traduce de {idioma_origen} a {idioma_destino} el siguiente contenido (no me des ninguna descripción extra del mismo) (si no te proporciono nada después de los dos puntos entonces imprime una x): {disciplina}.'
ordenes_intencion = f'Funge el rol de un {personalidad_modelo}. Ahora, traduce de {idioma_origen} a {idioma_destino} el siguiente contenido (si no te proporciono nada después de los dos puntos entonces imprime una x): {intencion}.'
ordenes_objetivo_general = f'Funge el rol de un {personalidad_modelo}. Ahora, traduce de {idioma_origen} a {idioma_destino} el siguiente contenido (si no te proporciono nada después de los dos puntos entonces imprime una x): {objetivo_general}.'
ordenes_temas = f'Funge el rol de un {personalidad_modelo}. Ahora, traduce de {idioma_origen} a {idioma_destino} el siguiente contenido (si no te proporciono nada después de los dos puntos entonces imprime una x): {temas}.'
ordenes_metodologia_ensenanza = f'Funge el rol de un {personalidad_modelo}. Ahora, traduce de {idioma_origen} a {idioma_destino} el siguiente contenido (si no te proporciono nada después de los dos puntos entonces imprime una x): {metodologia_ensenanza}.'
ordenes_tiempo_estimado = f'Funge el rol de un {personalidad_modelo}. Ahora, traduce de {idioma_origen} a {idioma_destino} el siguiente contenido (si no te proporciono nada después de los dos puntos entonces imprime una x): {tiempo_estimado}.'
ordenes_politica_evaluacion = f'Funge el rol de un {personalidad_modelo}. Ahora, traduce de {idioma_origen} a {idioma_destino} el siguiente contenido (si no te proporciono nada después de los dos puntos entonces imprime una x): {politica_evaluacion}.'
ordenes_perfil_profesor = f'Funge el rol de un {personalidad_modelo}. Ahora, traduce de {idioma_origen} a {idioma_destino} el siguiente contenido (si no te proporciono nada después de los dos puntos entonces imprime una x): {perfil_profesor}.'
ordenes_requisitos = f'Funge el rol de un {personalidad_modelo}. Ahora, traduce de {idioma_origen} a {idioma_destino} el siguiente contenido (si no te proporciono nada después de los dos puntos entonces imprime una x): {requisitos}.'
ordenes_descripcion = f'Funge el rol de un {personalidad_modelo}. Ahora, traduce de {idioma_origen} a {idioma_destino} el siguiente contenido (si no te proporciono nada después de los dos puntos entonces imprime una x): {descripcion}.'

#EXECUTE_THE_SKILL
resultado_nombre_materia = skpy.skill_exec_prompt(ordenes_nombre_materia, m_gpt_4)
resultado_disciplina = skpy.skill_exec_prompt(ordenes_disciplina, m_gpt_4)
resultado_intencion = skpy.skill_exec_prompt(ordenes_intencion, m_gpt_4)
resultado_objetivo_general = skpy.skill_exec_prompt(ordenes_objetivo_general, m_gpt_4)
resultado_temas = skpy.skill_exec_prompt(ordenes_temas, m_gpt_4)
resultado_metodologia_ensenanza = skpy.skill_exec_prompt(ordenes_metodologia_ensenanza, m_gpt_4)
resultado_tiempo_estimado = skpy.skill_exec_prompt(ordenes_tiempo_estimado, m_gpt_4)
resultado_politica_evaluacion = skpy.skill_exec_prompt(ordenes_politica_evaluacion, m_gpt_4)
resultado_perfil_profesor = skpy.skill_exec_prompt(ordenes_perfil_profesor, m_gpt_4)
resultado_requisitos = skpy.skill_exec_prompt(ordenes_requisitos, m_gpt_4)
resultado_descripcion = skpy.skill_exec_prompt(ordenes_descripcion, m_gpt_4)

#DEFINE_THE_OUTPUT
output = [resultado_nombre_materia, resultado_disciplina, resultado_intencion, resultado_objetivo_general, resultado_temas, resultado_metodologia_ensenanza, resultado_tiempo_estimado, resultado_politica_evaluacion, resultado_perfil_profesor, resultado_requisitos, resultado_descripcion]
'''

import openpyxl

def request_skrill(inputs):

    import requests
    import json

    url = 'https://tecgpt-grl-prod-skrill.azurewebsites.net/api/skrillpy_exec'
    params = {
        "id": 19,
        "email": "ezequiel.lozano@tec.mx",
        "inputs": [inputs],
        "code": codigo
    }

    json_params = json.dumps(params)
    response = requests.get(url, auth = ("LT09NGbinLZa1LI", "kB10Gsi96reJNQH"), data = json_params, headers = {'Content-Type': 'application/json'})

    if response.status_code == 200:
        return json.loads(response.text)
    else:
        return {
            "error": f"Request failed with status code: {response.status_code}"
        }

def procesar_excel(numero_filas):
    
    workbook = openpyxl.load_workbook('Ejemplo de tradución LAE19_sin traducción.xlsx')
    sheet = workbook['Sábanas de información']

    # Obtener las celdas fusionadas en la hoja de cálculo
    merged_cells = sheet.merged_cells.ranges
    inputs = []

    # Itera sobre las filas de la hoja de cálculo
    for row_index in range(4, numero_filas + 4):

        input = {}

        # Itera sobre las columnas de la hoja de cálculo
        for col_index in range(1, sheet.max_column + 1):
            # Obtiene el valor de la celda en la fila y columna actuales
            cell_value = sheet.cell(row=row_index, column=col_index).value
            # Verifica si la celda está fusionada
            is_merged = any(merged_cell.min_col <= col_index <= merged_cell.max_col and
                            merged_cell.min_row <= row_index <= merged_cell.max_row
                            for merged_cell in merged_cells)
            # Si la celda no está fusionada
            if not is_merged:
                if cell_value is not None:
                    if col_index == 2:
                        input["nombre_materia"] =  cell_value
                    if col_index == 4:
                        input["disciplina"] = cell_value
                    if col_index == 10:
                        input["intencion"] = cell_value
                    if col_index == 12:
                        input["objetivo_general"] = cell_value
                    if col_index == 14:
                        input["temas"] = cell_value
                    if col_index == 16:
                        input["metodologia_ensenanza"] = cell_value
                    if col_index == 18:
                        input["tiempo_estimado"] = cell_value
                    if col_index == 20:
                        input["politica_evaluacion"] = cell_value
                    if col_index == 22:
                        input["perfil_profesor"] = cell_value
                    if col_index == 24:
                        input["requisitos"] = cell_value
                    if col_index == 27:
                        input["descripcion"] = cell_value
        inputs.append(input)

    outputs = []
    
    for input in inputs:

        resultados_skrill = request_skrill(input)
        output = {}
        item_index = 1
        
        for resultado_modelo in resultados_skrill:
            if item_index == 1:
                output["nombre_materia"] =  resultado_modelo["result"]
            if item_index == 2:
                output["disciplina"] = resultado_modelo["result"]
            if item_index == 3:
                output["intencion"] = resultado_modelo["result"]
            if item_index == 4:
                output["objetivo_general"] = resultado_modelo["result"]
            if item_index == 5:
                output["temas"] = resultado_modelo["result"]
            if item_index == 6:
                output["metodologia_ensenanza"] = resultado_modelo["result"]
            if item_index == 7:
                output["tiempo_estimado"] = resultado_modelo["result"]
            if item_index == 8:
                output["politica_evaluacion"] = resultado_modelo["result"]
            if item_index == 9:
                output["perfil_profesor"] = resultado_modelo["result"]
            if item_index == 10:
                output["requisitos"] = resultado_modelo["result"]
            if item_index == 11:
                output["descripcion"] = resultado_modelo["result"]
            item_index += 1
        outputs.append(output)
        
    item_outputs = 0

    # Itera sobre las filas de la hoja de cálculo
    for row_index in range(4, numero_filas + 4):

        # Itera sobre las columnas de la hoja de cálculo
        for col_index in range(1, sheet.max_column + 1):
            # Obtiene el valor de la celda en la fila y columna actuales
            cell_value = sheet.cell(row=row_index, column=col_index).value
            # Verifica si la celda está fusionada
            is_merged = any(merged_cell.min_col <= col_index <= merged_cell.max_col and
                            merged_cell.min_row <= row_index <= merged_cell.max_row
                            for merged_cell in merged_cells)
            # Si la celda está fusionada
            if not is_merged:
                # Si la celda está vacía
                if cell_value is None:
                    if col_index == 3:
                        cell_value = outputs[item_outputs]["nombre_materia"]
                    if col_index == 5:
                        cell_value = outputs[item_outputs]["disciplina"]
                    if col_index == 11:
                        cell_value = outputs[item_outputs]["intencion"]
                    if col_index == 13:
                        cell_value = outputs[item_outputs]["objetivo_general"]
                    if col_index == 15:
                        cell_value = outputs[item_outputs]["temas"]
                    if col_index == 17:
                        cell_value = outputs[item_outputs]["metodologia_ensenanza"]
                    if col_index == 19:
                        cell_value = outputs[item_outputs]["tiempo_estimado"]
                    if col_index == 21:
                        cell_value = outputs[item_outputs]["politica_evaluacion"]
                    if col_index == 23:
                        cell_value = outputs[item_outputs]["perfil_profesor"]
                    if col_index == 25:
                        cell_value = outputs[item_outputs]["requisitos"]
                    if col_index == 28:
                        cell_value = outputs[item_outputs]["descripcion"]
                    sheet.cell(row=row_index, column=col_index).value = cell_value
        item_outputs += 1
    # Guarda los cambios en un nuevo archivo Excel
    workbook.save('Traducción LAE19 - DEMO.xlsx')

procesar_excel(127)